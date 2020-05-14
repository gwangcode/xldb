#! /Library/Frameworks/Python.framework/Versions/3.6/bin/python3

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from styledtext import grab_property, write_property, bool_str, grab_text, write_text
from datatype import str2data
import sys, re, stdwb as wb

# font=Calibri size=11 mode=bold+italic+undrline+strike+baseline+superscript+subscript forecolor=FF0000000 backcolor=FFFFFFFF 
# bleft=solid FF000000 brigt=solid FF000000 btop=solid FF000000 bbottom=solid FF000000 bdiagonal=solid FF000000 0 boutline=solid FF000000 bhorizontal=solid FF000000 bvertical=solid FF000000 
# alhorizontal=general alvertical=bottom rotation=0 indent=0 alignment=wrap fit 
# format=General 
# protection=lock hide

def write_style(CellObj, s):
  bold=False
  italic=False
  vertAlign=None
  underline='none'
  strike=False

  name=grab_property('font', s)
  if not name: name='Calibri'
  size=grab_property('size', s)
  if not size: size=11
  mode=grab_property('mode', s)
  if mode:
    for m in mode.split('+'):
      m=m.strip()
      if m=='bold': bold=True #bold
      elif m=='italic': italic=True 
      elif m=='underline_double': underline='double'
      elif m=='underline_double_accounting': underline='doubleAccounting'
      elif m=='underline_single': underline='single'
      elif m=='underline_single_accounting': underline='singleAccounting'
      elif m=='strike': strike=True
      elif m=='base': vertAlign='base'
      elif m=='superscript': vertAlign='superscript'
      elif m=='subscript': vertAlign='subscript'

  forecolor=grab_property('forecolor', s)
  if not forecolor: forecolor='FF000000'
  filltype=grab_property('filltype', s)
  backcolor=grab_property('backcolor', s)
  if not backcolor: 
    backcolor='FFFFFFFF'
    filltype=None
  

  a=('bleft', 'bright', 'btop', 'bbottom', 'bdiagonal', 'bhorizontal', 'bvertical')
  bd=[]
  for i, n in zip(a, range(len(a))):
    p=grab_property(i, s)
    if p: 
      b=p.split('+')
      if len(b)==1: side=Side(border_style=b[0])
      elif len(b)==2: side=Side(border_style=b[0], color=b[1])
      else: side=Side()
    else: side=Side()
    bd.append(side)
  
  boutline=grab_property('boutline', s)
  if boutline: boutline=bool_str(s)
  else: boutline=False


  alhorizontal=grab_property('alhorizontal', s)
  if not alhorizontal: alhorizontal='general'

  alvertical=grab_property('alvertical', s)
  if not alhorizontal: alvertical='bottom'

  rotation=grab_property('rotation', s)
  if rotation: rotation=int(rotation)
  else: rotation=0

  indent=grab_property('indent', s)
  if indent: rotation=int(indent)
  else: indent=0

  alignment=grab_property('alignment', s)
  wrap=False
  fit=False
  if alignment:
    for p in alignment.split('+'):
      p=p.strip()
      if p=='wrap': wrap==True
      elif p=='fit': fit=True

  num_format=grab_property('format', s)
  if not num_format: num_format='General'

  protection=grab_property('protection', s)
  locked=False
  hidden=False
  if alignment:
    for p in protection.split('+'):
      p=p.strip()
      if p=='lock': locked==True
      elif p=='hide': hidden=True
  
  CellObj.font = Font(name=name, size=size, bold=bold, italic=italic, vertAlign=vertAlign, underline=underline, strike=strike, color=forecolor)
  CellObj.fill = PatternFill(fill_type=filltype, bgColor=backcolor)
  CellObj.border = Border(left=bd[0], right=bd[1], top=bd[2], bottom=bd[3], diagonal=bd[4], diagonal_direction=0, outline=boutline, horizontal=bd[5], vertical=bd[6])
  CellObj.alignment=Alignment(horizontal=alhorizontal, vertical=alvertical, text_rotation=rotation, wrap_text=wrap, shrink_to_fit=fit, indent=indent)
  CellObj.protection = Protection(locked=locked, hidden=hidden)
  CellObj.number_format = num_format
  
def read_style(CellObj):
  if CellObj.font:
    name=CellObj.font.name
    size=CellObj.font.size
    bold=CellObj.font.bold
    italic=CellObj.font.italic
    vertAlign=CellObj.font.vertAlign
    underline=CellObj.font.underline
    strike=CellObj.font.strike
    forecolor=CellObj.font.color
    if forecolor: forecolor=forecolor.rgb
    else: forecolor=None
  else:
    name=None
    size=None
    bold=None
    italic=None
    vertAlign=None
    underline=None
    strike=None
    forecolor=None
  if CellObj.fill: 
    filltype=CellObj.fill.fill_type
    backcolor=CellObj.fill.bgColor
    #backcolor=CellObj.fill.start_color
    if backcolor: backcolor=backcolor.rgb
    else: 
      backcolor=None
      filltype=None
  else: 
    backcolor=None
    filltype=None
  if CellObj.border.left: bleft=(CellObj.border.left.style, CellObj.border.left.color)
  else: bleft=None
  if CellObj.border.right: bright=(CellObj.border.right.style, CellObj.border.right.color)
  else: bright=None
  if CellObj.border.top: btop=(CellObj.border.top.style, CellObj.border.top.color)
  else: btop=None
  if CellObj.border.bottom: bbottom=(CellObj.border.bottom.style, CellObj.border.bottom.color)
  else: bbottom=None
  if CellObj.border.diagonal: bdiagonal=(CellObj.border.diagonal.style, CellObj.border.diagonal.color)
  else: bdiagonal=None 
  boutline=CellObj.border.outline
  if CellObj.border.vertical: bvertical=(CellObj.border.vertical.style, CellObj.border.vertical.color)
  else: bvertical=None
  if CellObj.border.horizontal: bhorizontal=(CellObj.border.horizontal.style, CellObj.border.horizontal.color)
  else: bhorizontal=None
  if CellObj.alignment:
    alhorizontal=CellObj.alignment.horizontal
    alvertical=CellObj.alignment.vertical
    rotation=CellObj.alignment.text_rotation
    wrap=CellObj.alignment.wrap_text
    fit=CellObj.alignment.shrink_to_fit
    indent=CellObj.alignment.indent
  else:
    alhorizontal=None
    alvertical=None
    rotation=None
    wrap=None
    fit=None
    indent=None
  if CellObj.protection:
    locked=CellObj.protection.locked
    hidden=CellObj.protection.hidden
  else:
    locked=None
    hidden=None
  num_format=CellObj.number_format

# format=General 
# protection=lock hide
  s='<|>'
  if name: s=write_property('font', name, s)
  if size is not None: s=write_property('size', size, s)
  mode=''
  if bold: mode='bold'
  if mode:
    if italic: mode+='+italic'
  else:
    if italic: mode='italic'
  if mode:
    if underline: mode+='+'+underline
  else:
    if underline: mode=underline
  if mode:
    if strike: mode+='+strike'
  else:
    if strike: mode='strike'
  if mode:
    if vertAlign: mode+='+'+vertAlign
  else:
    if vertAlign: mode=vertAlign
  if mode: s=write_property('mode', mode, s)

  if forecolor: s=write_property('forecolor', forecolor, s)
  if backcolor: s=write_property('backcolor', backcolor, s)
  if filltype: s=write_property('filltype', filltype, s)
  
  a=('bleft', 'bright', 'btop', 'bbottom', 'bdiagonal', 'bhorizontal', 'bvertical')
  b=(bleft, bright, btop, bbottom, bdiagonal, bhorizontal, bvertical)
  for i, j in zip(a,b):
    if j: 
      if j[0]: 
        s=write_property(i, j[0], s)
        if j[1]: s=write_property(i, j[0]+'+'+j[1], s)
      elif j[1]: s=write_property(i, j[1], s)

  if boutline: s=write_property('boutline', boutline, s)
  
  a=('alhorizontal', 'alvertical', 'rotation', 'indent')
  b=(alhorizontal, alvertical, rotation, indent)
  for i, j in zip(a,b): 
    if j: s=write_property(i, j, s)
      
  p=''
  if wrap: 
    p='wrap'
    if fit: p='wrap+fit'
  elif fit: p='fit'
  if p: s=write_property('alignment', p, s)
  if num_format: s=write_property('format', num_format, s)
  p=''
  if locked: 
    p='lock'
    if hidden: p='lock+hide'
  elif hidden: p='hide'
  if p: s=write_property('protection', p, s)
  p=wb.num2col(CellObj.column)+str(CellObj.row) # cell pos: A2
  if p: s=write_property('cell', p, s)

  return s

CLUT = [  # color look-up table
#    8-bit, RGB hex

    # Primary 3-bit (8 colors). Unique representation!
    ('0',  '000000'),
    ('1',  '800000'),
    ('2',  '008000'),
    ('3',  '808000'),
    ('4',  '000080'),
    ('5',  '800080'),
    ('6',  '008080'),
    ('7',  'c0c0c0'),

    # Equivalent "bright" versions of original 8 colors.
    ('8',  '808080'),
    ('9',  'ff0000'),
    ('10',  '00ff00'),
    ('11',  'ffff00'),
    ('12',  '0000ff'),
    ('13',  'ff00ff'),
    ('14',  '00ffff'),
    ('15',  'ffffff'),

    # Strictly ascending.
    ('16',  '000000'),
    ('17',  '00005f'),
    ('18',  '000087'),
    ('19',  '0000af'),
    ('20',  '0000d7'),
    ('21',  '0000ff'),
    ('22',  '005f00'),
    ('23',  '005f5f'),
    ('24',  '005f87'),
    ('25',  '005faf'),
    ('26',  '005fd7'),
    ('27',  '005fff'),
    ('28',  '008700'),
    ('29',  '00875f'),
    ('30',  '008787'),
    ('31',  '0087af'),
    ('32',  '0087d7'),
    ('33',  '0087ff'),
    ('34',  '00af00'),
    ('35',  '00af5f'),
    ('36',  '00af87'),
    ('37',  '00afaf'),
    ('38',  '00afd7'),
    ('39',  '00afff'),
    ('40',  '00d700'),
    ('41',  '00d75f'),
    ('42',  '00d787'),
    ('43',  '00d7af'),
    ('44',  '00d7d7'),
    ('45',  '00d7ff'),
    ('46',  '00ff00'),
    ('47',  '00ff5f'),
    ('48',  '00ff87'),
    ('49',  '00ffaf'),
    ('50',  '00ffd7'),
    ('51',  '00ffff'),
    ('52',  '5f0000'),
    ('53',  '5f005f'),
    ('54',  '5f0087'),
    ('55',  '5f00af'),
    ('56',  '5f00d7'),
    ('57',  '5f00ff'),
    ('58',  '5f5f00'),
    ('59',  '5f5f5f'),
    ('60',  '5f5f87'),
    ('61',  '5f5faf'),
    ('62',  '5f5fd7'),
    ('63',  '5f5fff'),
    ('64',  '5f8700'),
    ('65',  '5f875f'),
    ('66',  '5f8787'),
    ('67',  '5f87af'),
    ('68',  '5f87d7'),
    ('69',  '5f87ff'),
    ('70',  '5faf00'),
    ('71',  '5faf5f'),
    ('72',  '5faf87'),
    ('73',  '5fafaf'),
    ('74',  '5fafd7'),
    ('75',  '5fafff'),
    ('76',  '5fd700'),
    ('77',  '5fd75f'),
    ('78',  '5fd787'),
    ('79',  '5fd7af'),
    ('80',  '5fd7d7'),
    ('81',  '5fd7ff'),
    ('82',  '5fff00'),
    ('83',  '5fff5f'),
    ('84',  '5fff87'),
    ('85',  '5fffaf'),
    ('86',  '5fffd7'),
    ('87',  '5fffff'),
    ('88',  '870000'),
    ('89',  '87005f'),
    ('90',  '870087'),
    ('91',  '8700af'),
    ('92',  '8700d7'),
    ('93',  '8700ff'),
    ('94',  '875f00'),
    ('95',  '875f5f'),
    ('96',  '875f87'),
    ('97',  '875faf'),
    ('98',  '875fd7'),
    ('99',  '875fff'),
    ('100', '878700'),
    ('101', '87875f'),
    ('102', '878787'),
    ('103', '8787af'),
    ('104', '8787d7'),
    ('105', '8787ff'),
    ('106', '87af00'),
    ('107', '87af5f'),
    ('108', '87af87'),
    ('109', '87afaf'),
    ('110', '87afd7'),
    ('111', '87afff'),
    ('112', '87d700'),
    ('113', '87d75f'),
    ('114', '87d787'),
    ('115', '87d7af'),
    ('116', '87d7d7'),
    ('117', '87d7ff'),
    ('118', '87ff00'),
    ('119', '87ff5f'),
    ('120', '87ff87'),
    ('121', '87ffaf'),
    ('122', '87ffd7'),
    ('123', '87ffff'),
    ('124', 'af0000'),
    ('125', 'af005f'),
    ('126', 'af0087'),
    ('127', 'af00af'),
    ('128', 'af00d7'),
    ('129', 'af00ff'),
    ('130', 'af5f00'),
    ('131', 'af5f5f'),
    ('132', 'af5f87'),
    ('133', 'af5faf'),
    ('134', 'af5fd7'),
    ('135', 'af5fff'),
    ('136', 'af8700'),
    ('137', 'af875f'),
    ('138', 'af8787'),
    ('139', 'af87af'),
    ('140', 'af87d7'),
    ('141', 'af87ff'),
    ('142', 'afaf00'),
    ('143', 'afaf5f'),
    ('144', 'afaf87'),
    ('145', 'afafaf'),
    ('146', 'afafd7'),
    ('147', 'afafff'),
    ('148', 'afd700'),
    ('149', 'afd75f'),
    ('150', 'afd787'),
    ('151', 'afd7af'),
    ('152', 'afd7d7'),
    ('153', 'afd7ff'),
    ('154', 'afff00'),
    ('155', 'afff5f'),
    ('156', 'afff87'),
    ('157', 'afffaf'),
    ('158', 'afffd7'),
    ('159', 'afffff'),
    ('160', 'd70000'),
    ('161', 'd7005f'),
    ('162', 'd70087'),
    ('163', 'd700af'),
    ('164', 'd700d7'),
    ('165', 'd700ff'),
    ('166', 'd75f00'),
    ('167', 'd75f5f'),
    ('168', 'd75f87'),
    ('169', 'd75faf'),
    ('170', 'd75fd7'),
    ('171', 'd75fff'),
    ('172', 'd78700'),
    ('173', 'd7875f'),
    ('174', 'd78787'),
    ('175', 'd787af'),
    ('176', 'd787d7'),
    ('177', 'd787ff'),
    ('178', 'd7af00'),
    ('179', 'd7af5f'),
    ('180', 'd7af87'),
    ('181', 'd7afaf'),
    ('182', 'd7afd7'),
    ('183', 'd7afff'),
    ('184', 'd7d700'),
    ('185', 'd7d75f'),
    ('186', 'd7d787'),
    ('187', 'd7d7af'),
    ('188', 'd7d7d7'),
    ('189', 'd7d7ff'),
    ('190', 'd7ff00'),
    ('191', 'd7ff5f'),
    ('192', 'd7ff87'),
    ('193', 'd7ffaf'),
    ('194', 'd7ffd7'),
    ('195', 'd7ffff'),
    ('196', 'ff0000'),
    ('197', 'ff005f'),
    ('198', 'ff0087'),
    ('199', 'ff00af'),
    ('200', 'ff00d7'),
    ('201', 'ff00ff'),
    ('202', 'ff5f00'),
    ('203', 'ff5f5f'),
    ('204', 'ff5f87'),
    ('205', 'ff5faf'),
    ('206', 'ff5fd7'),
    ('207', 'ff5fff'),
    ('208', 'ff8700'),
    ('209', 'ff875f'),
    ('210', 'ff8787'),
    ('211', 'ff87af'),
    ('212', 'ff87d7'),
    ('213', 'ff87ff'),
    ('214', 'ffaf00'),
    ('215', 'ffaf5f'),
    ('216', 'ffaf87'),
    ('217', 'ffafaf'),
    ('218', 'ffafd7'),
    ('219', 'ffafff'),
    ('220', 'ffd700'),
    ('221', 'ffd75f'),
    ('222', 'ffd787'),
    ('223', 'ffd7af'),
    ('224', 'ffd7d7'),
    ('225', 'ffd7ff'),
    ('226', 'ffff00'),
    ('227', 'ffff5f'),
    ('228', 'ffff87'),
    ('229', 'ffffaf'),
    ('230', 'ffffd7'),
    ('231', 'ffffff'),

    # Gray-scale range.
    ('232', '080808'),
    ('233', '121212'),
    ('234', '1c1c1c'),
    ('235', '262626'),
    ('236', '303030'),
    ('237', '3a3a3a'),
    ('238', '444444'),
    ('239', '4e4e4e'),
    ('240', '585858'),
    ('241', '626262'),
    ('242', '6c6c6c'),
    ('243', '767676'),
    ('244', '808080'),
    ('245', '8a8a8a'),
    ('246', '949494'),
    ('247', '9e9e9e'),
    ('248', 'a8a8a8'),
    ('249', 'b2b2b2'),
    ('250', 'bcbcbc'),
    ('251', 'c6c6c6'),
    ('252', 'd0d0d0'),
    ('253', 'dadada'),
    ('254', 'e4e4e4'),
    ('255', 'eeeeee'),
]

# 256 colors name & color exchange chart
Color2Name={0: 'Black', 1: 'Maroon', 2: 'Green', 3: 'Olive', 4: 'Navy', 5: 'Purple', 6: 'Teal', 7: 'Silver', 8: 'Grey', 9: 'Red', 10: 'Lime', 11: 'Yellow', 12: 'Blue', 13: 'Fuchsia', 14: 'Aqua', 15: 'White', 16: 'Grey0', 17: 'NavyBlue', 18: 'DarkBlue', 19: 'Blue3', 20: 'Blue3', 21: 'Blue1', 22: 'DarkGreen', 23: 'DeepSkyBlue4', 24: 'DeepSkyBlue4', 25: 'DeepSkyBlue4', 26: 'DodgerBlue3', 27: 'DodgerBlue2', 28: 'Green4', 29: 'SpringGreen4', 30: 'Turquoise4', 31: 'DeepSkyBlue3', 32: 'DeepSkyBlue3', 33: 'DodgerBlue1', 34: 'Green3', 35: 'SpringGreen3', 36: 'DarkCyan', 37: 'LightSeaGreen', 38: 'DeepSkyBlue2', 39: 'DeepSkyBlue1', 40: 'Green3', 41: 'SpringGreen3', 42: 'SpringGreen2', 43: 'Cyan3', 44: 'DarkTurquoise', 45: 'Turquoise2', 46: 'Green1', 47: 'SpringGreen2', 48: 'SpringGreen1', 49: 'MediumSpringGreen', 50: 'Cyan2', 51: 'Cyan1', 52: 'DarkRed', 53: 'DeepPink4', 54: 'Purple4', 55: 'Purple4', 56: 'Purple3', 57: 'BlueViolet', 58: 'Orange4', 59: 'Grey37', 60: 'MediumPurple4', 61: 'SlateBlue3', 62: 'SlateBlue3', 63: 'RoyalBlue1', 64: 'Chartreuse4', 65: 'DarkSeaGreen4', 66: 'PaleTurquoise4', 67: 'SteelBlue', 68: 'SteelBlue3', 69: 'CornflowerBlue', 70: 'Chartreuse3', 71: 'DarkSeaGreen4', 72: 'CadetBlue', 73: 'CadetBlue', 74: 'SkyBlue3', 75: 'SteelBlue1', 76: 'Chartreuse3', 77: 'PaleGreen3', 78: 'SeaGreen3', 79: 'Aquamarine3', 80: 'MediumTurquoise', 81: 'SteelBlue1', 82: 'Chartreuse2', 83: 'SeaGreen2', 84: 'SeaGreen1', 85: 'SeaGreen1', 86: 'Aquamarine1', 87: 'DarkSlateGray2', 88: 'DarkRed', 89: 'DeepPink4', 90: 'DarkMagenta', 91: 'DarkMagenta', 92: 'DarkViolet', 93: 'Purple', 94: 'Orange4', 95: 'LightPink4', 96: 'Plum4', 97: 'MediumPurple3', 98: 'MediumPurple3', 99: 'SlateBlue1', 100: 'Yellow4', 101: 'Wheat4', 102: 'Grey53', 103: 'LightSlateGrey', 104: 'MediumPurple', 105: 'LightSlateBlue', 106: 'Yellow4', 107: 'DarkOliveGreen3', 108: 'DarkSeaGreen', 109: 'LightSkyBlue3', 110: 'LightSkyBlue3', 111: 'SkyBlue2', 112: 'Chartreuse2', 113: 'DarkOliveGreen3', 114: 'PaleGreen3', 115: 'DarkSeaGreen3', 116: 'DarkSlateGray3', 117: 'SkyBlue1', 118: 'Chartreuse1', 119: 'LightGreen', 120: 'LightGreen', 121: 'PaleGreen1', 122: 'Aquamarine1', 123: 'DarkSlateGray1', 124: 'Red3', 125: 'DeepPink4', 126: 'MediumVioletRed', 127: 'Magenta3', 128: 'DarkViolet', 129: 'Purple', 130: 'DarkOrange3', 131: 'IndianRed', 132: 'HotPink3', 133: 'MediumOrchid3', 134: 'MediumOrchid', 135: 'MediumPurple2', 136: 'DarkGoldenrod', 137: 'LightSalmon3', 138: 'RosyBrown', 139: 'Grey63', 140: 'MediumPurple2', 141: 'MediumPurple1', 142: 'Gold3', 143: 'DarkKhaki', 144: 'NavajoWhite3', 145: 'Grey69', 146: 'LightSteelBlue3', 147: 'LightSteelBlue', 148: 'Yellow3', 149: 'DarkOliveGreen3', 150: 'DarkSeaGreen3', 151: 'DarkSeaGreen2', 152: 'LightCyan3', 153: 'LightSkyBlue1', 154: 'GreenYellow', 155: 'DarkOliveGreen2', 156: 'PaleGreen1', 157: 'DarkSeaGreen2', 158: 'DarkSeaGreen1', 159: 'PaleTurquoise1', 160: 'Red3', 161: 'DeepPink3', 162: 'DeepPink3', 163: 'Magenta3', 164: 'Magenta3', 165: 'Magenta2', 166: 'DarkOrange3', 167: 'IndianRed', 168: 'HotPink3', 169: 'HotPink2', 170: 'Orchid', 171: 'MediumOrchid1', 172: 'Orange3', 173: 'LightSalmon3', 174: 'LightPink3', 175: 'Pink3', 176: 'Plum3', 177: 'Violet', 178: 'Gold3', 179: 'LightGoldenrod3', 180: 'Tan', 181: 'MistyRose3', 182: 'Thistle3', 183: 'Plum2', 184: 'Yellow3', 185: 'Khaki3', 186: 'LightGoldenrod2', 187: 'LightYellow3', 188: 'Grey84', 189: 'LightSteelBlue1', 190: 'Yellow2', 191: 'DarkOliveGreen1', 192: 'DarkOliveGreen1', 193: 'DarkSeaGreen1', 194: 'Honeydew2', 195: 'LightCyan1', 196: 'Red1', 197: 'DeepPink2', 198: 'DeepPink1', 199: 'DeepPink1', 200: 'Magenta2', 201: 'Magenta1', 202: 'OrangeRed1', 203: 'IndianRed1', 204: 'IndianRed1', 205: 'HotPink', 206: 'HotPink', 207: 'MediumOrchid1', 208: 'DarkOrange', 209: 'Salmon1', 210: 'LightCoral', 211: 'PaleVioletRed1', 212: 'Orchid2', 213: 'Orchid1', 214: 'Orange1', 215: 'SandyBrown', 216: 'LightSalmon1', 217: 'LightPink1', 218: 'Pink1', 219: 'Plum1', 220: 'Gold1', 221: 'LightGoldenrod2', 222: 'LightGoldenrod2', 223: 'NavajoWhite1', 224: 'MistyRose1', 225: 'Thistle1', 226: 'Yellow1', 227: 'LightGoldenrod1', 228: 'Khaki1', 229: 'Wheat1', 230: 'Cornsilk1', 231: 'Grey100', 232: 'Grey3', 233: 'Grey7', 234: 'Grey11', 235: 'Grey15', 236: 'Grey19', 237: 'Grey23', 238: 'Grey27', 239: 'Grey30', 240: 'Grey35', 241: 'Grey39', 242: 'Grey42', 243: 'Grey46', 244: 'Grey50', 245: 'Grey54', 246: 'Grey58', 247: 'Grey62', 248: 'Grey66', 249: 'Grey70', 250: 'Grey74', 251: 'Grey78', 252: 'Grey82', 253: 'Grey85', 254: 'Grey89', 255: 'Grey93'}
Name2Color={'Black': 0, 'Maroon': 1, 'Green': 2, 'Olive': 3, 'Navy': 4, 'Purple': 129, 'Teal': 6, 'Silver': 7, 'Grey': 8, 'Red': 9, 'Lime': 10, 'Yellow': 11, 'Blue': 12, 'Fuchsia': 13, 'Aqua': 14, 'White': 15, 'Grey0': 16, 'NavyBlue': 17, 'DarkBlue': 18, 'Blue3': 20, 'Blue1': 21, 'DarkGreen': 22, 'DeepSkyBlue4': 25, 'DodgerBlue3': 26, 'DodgerBlue2': 27, 'Green4': 28, 'SpringGreen4': 29, 'Turquoise4': 30, 'DeepSkyBlue3': 32, 'DodgerBlue1': 33, 'Green3': 40, 'SpringGreen3': 41, 'DarkCyan': 36, 'LightSeaGreen': 37, 'DeepSkyBlue2': 38, 'DeepSkyBlue1': 39, 'SpringGreen2': 47, 'Cyan3': 43, 'DarkTurquoise': 44, 'Turquoise2': 45, 'Green1': 46, 'SpringGreen1': 48, 'MediumSpringGreen': 49, 'Cyan2': 50, 'Cyan1': 51, 'DarkRed': 88, 'DeepPink4': 125, 'Purple4': 55, 'Purple3': 56, 'BlueViolet': 57, 'Orange4': 94, 'Grey37': 59, 'MediumPurple4': 60, 'SlateBlue3': 62, 'RoyalBlue1': 63, 'Chartreuse4': 64, 'DarkSeaGreen4': 71, 'PaleTurquoise4': 66, 'SteelBlue': 67, 'SteelBlue3': 68, 'CornflowerBlue': 69, 'Chartreuse3': 76, 'CadetBlue': 73, 'SkyBlue3': 74, 'SteelBlue1': 81, 'PaleGreen3': 114, 'SeaGreen3': 78, 'Aquamarine3': 79, 'MediumTurquoise': 80, 'Chartreuse2': 112, 'SeaGreen2': 83, 'SeaGreen1': 85, 'Aquamarine1': 122, 'DarkSlateGray2': 87, 'DarkMagenta': 91, 'DarkViolet': 128, 'LightPink4': 95, 'Plum4': 96, 'MediumPurple3': 98, 'SlateBlue1': 99, 'Yellow4': 106, 'Wheat4': 101, 'Grey53': 102, 'LightSlateGrey': 103, 'MediumPurple': 104, 'LightSlateBlue': 105, 'DarkOliveGreen3': 149, 'DarkSeaGreen': 108, 'LightSkyBlue3': 110, 'SkyBlue2': 111, 'DarkSeaGreen3': 150, 'DarkSlateGray3': 116, 'SkyBlue1': 117, 'Chartreuse1': 118, 'LightGreen': 120, 'PaleGreen1': 156, 'DarkSlateGray1': 123, 'Red3': 160, 'MediumVioletRed': 126, 'Magenta3': 164, 'DarkOrange3': 166, 'IndianRed': 167, 'HotPink3': 168, 'MediumOrchid3': 133, 'MediumOrchid': 134, 'MediumPurple2': 140, 'DarkGoldenrod': 136, 'LightSalmon3': 173, 'RosyBrown': 138, 'Grey63': 139, 'MediumPurple1': 141, 'Gold3': 178, 'DarkKhaki': 143, 'NavajoWhite3': 144, 'Grey69': 145, 'LightSteelBlue3': 146, 'LightSteelBlue': 147, 'Yellow3': 184, 'DarkSeaGreen2': 157, 'LightCyan3': 152, 'LightSkyBlue1': 153, 'GreenYellow': 154, 'DarkOliveGreen2': 155, 'DarkSeaGreen1': 193, 'PaleTurquoise1': 159, 'DeepPink3': 162, 'Magenta2': 200, 'HotPink2': 169, 'Orchid': 170, 'MediumOrchid1': 207, 'Orange3': 172, 'LightPink3': 174, 'Pink3': 175, 'Plum3': 176, 'Violet': 177, 'LightGoldenrod3': 179, 'Tan': 180, 'MistyRose3': 181, 'Thistle3': 182, 'Plum2': 183, 'Khaki3': 185, 'LightGoldenrod2': 222, 'LightYellow3': 187, 'Grey84': 188, 'LightSteelBlue1': 189, 'Yellow2': 190, 'DarkOliveGreen1': 192, 'Honeydew2': 194, 'LightCyan1': 195, 'Red1': 196, 'DeepPink2': 197, 'DeepPink1': 199, 'Magenta1': 201, 'OrangeRed1': 202, 'IndianRed1': 204, 'HotPink': 206, 'DarkOrange': 208, 'Salmon1': 209, 'LightCoral': 210, 'PaleVioletRed1': 211, 'Orchid2': 212, 'Orchid1': 213, 'Orange1': 214, 'SandyBrown': 215, 'LightSalmon1': 216, 'LightPink1': 217, 'Pink1': 218, 'Plum1': 219, 'Gold1': 220, 'NavajoWhite1': 223, 'MistyRose1': 224, 'Thistle1': 225, 'Yellow1': 226, 'LightGoldenrod1': 227, 'Khaki1': 228, 'Wheat1': 229, 'Cornsilk1': 230, 'Grey100': 231, 'Grey3': 232, 'Grey7': 233, 'Grey11': 234, 'Grey15': 235, 'Grey19': 236, 'Grey23': 237, 'Grey27': 238, 'Grey30': 239, 'Grey35': 240, 'Grey39': 241, 'Grey42': 242, 'Grey46': 243, 'Grey50': 244, 'Grey54': 245, 'Grey58': 246, 'Grey62': 247, 'Grey66': 248, 'Grey70': 249, 'Grey74': 250, 'Grey78': 251, 'Grey82': 252, 'Grey85': 253, 'Grey89': 254, 'Grey93': 255}

def _str2hex(hexstr):
    return int(hexstr, 16)

def _strip_hash(rgb):
    # Strip leading `#` if exists.
    if rgb.startswith('#'):
        rgb = rgb.lstrip('#')
    return rgb

def _create_dicts():
    short2rgb_dict = dict(CLUT)
    rgb2short_dict = {}
    for k, v in short2rgb_dict.items():
        rgb2short_dict[v] = k
    return rgb2short_dict, short2rgb_dict

RGB2SHORT_DICT, SHORT2RGB_DICT = _create_dicts()

# 256 to RGB
# transparency: 0-1
def short2rgb(short, transparency=None):
  r=SHORT2RGB_DICT[short]
  if transparency: 
    h=hex(int(transparency*255))[2:]
    if len(h)==1: h='0'+h
    r=h+r
  return r

# RGB to 256
def rgb2short(rgb):
    """ Find the closest xterm-256 approximation to the given RGB value.
    @param rgb: Hex code representing an RGB value, eg, 'abcdef'
    @returns: String between 0 and 255, compatible with xterm.
    default_rgb2short('123456')
    ('23', '005f5f')
    default_rgb2short('ffffff')
    ('231', 'ffffff')
    default_rgb2short('0DADD6') # vimeo logo
    ('38', '00afd7')
    """
    rgb=rgb[2:]
    rgb = _strip_hash(rgb)
    incs = (0x00, 0x5f, 0x87, 0xaf, 0xd7, 0xff)
    # Break 6-char RGB code into 3 integer vals.
    parts = [ int(h, 16) for h in re.split(r'(..)(..)(..)', rgb)[1:4] ]
    res = []
    for part in parts:
        i = 0
        while i < len(incs)-1:
            s, b = incs[i], incs[i+1]  # smaller, bigger
            if s <= part <= b:
                s1 = abs(s - part)
                b1 = abs(b - part)
                if s1 < b1: closest = s
                else: closest = b
                res.append(closest)
                break
            i += 1
    #print '***', res
    res = ''.join([ ('%02.x' % i) for i in res ])
    equiv = RGB2SHORT_DICT[ res ]
    #print '***', res, equiv
    return equiv

# font=Calibri size=11 mode=bold+italic+underline+strike+baseline+superscript+subscript forecolor=FF0000000 backcolor=FFFFFFFF 
# bleft=solid FF000000 brigth=solid FF000000 btop=solid FF000000 bbottom=solid FF000000 bdiagonal=solid FF000000 0 boutline=solid FF000000 bhorizontal=solid FF000000 bvertical=solid FF000000 
# alhorizontal=general alvertical=bottom rotation=0 indent=0 alignment=wrap fit 
# format=General 
# protection=lock hide

# Normal2Console
def mode_normal2console(attributes):
  mode=''
  modes=grab_property('mode', attributes)
  if modes:
    modes=modes.split('+')
    if 'bold' in modes: mode='bold' # bold

    if 'italic' in modes: # italic
      if mode: mode+='+italic'
      else: mode='italic'

    if 'underline' in modes: # underline
      if mode: mode+='+underline'
      else: mode='underline'
  
  return mode

# <line=0 col=0 forecolor=2 backcolor=5 mode=UNDERLINE|Hello world!>
# <font= bleft= |Hello world> => mode=bold+underline forecolor=15 backcolor=30
def normal2console(bracket):
  print(bracket)
  r='<|>'
  init_mode=grab_property('mode', bracket)
  if init_mode: 
    mode=mode_normal2console(init_mode)
    r=write_property('mode', mode, r)

  cell=grab_property('cell', bracket)
  if cell: r=write_property('cell', cell, r)

  init_forecolor=grab_property('forecolor',bracket)
  if init_forecolor: 
    forecolor=rgb2short(init_forecolor)
    r=write_property('forecolor', forecolor, r)
  
  init_backcolor=grab_property('backcolor',bracket)
  filltype=grab_property('filltype',bracket)
  if init_backcolor: 
    backcolor=rgb2short(init_backcolor)
    if filltype: r=write_property('backcolor', backcolor, r)
    else: r=write_property('backcolor', 231, r)
    
  text=grab_text(bracket)
  if text:
    r=write_text(text, r)

  return r


# Console2Normal
def console2normal(bracket):
  forecolor=short2rgb(grab_property('forecolor',bracket), transparency=1)
  backcolor=short2rgb(grab_property('backcolor',bracket), transparency=1)
  bracket=write_property('forecolor', forecolor, bracket)
  if backcolor:
    bracket=write_property('backcolor', backcolor, bracket)
    bracket=write_property('filltype', 'solid', bracket)
  return bracket


#Mode=Normal: normal original style, Console: 256 color & curses style
# read cell as <style|value>
def cread(CellObj, Mode='Normal'):
  bracket=read_style(CellObj)
  v=CellObj.value
  if v: v=str(v)
  else: v=''
  r=write_text(v, bracket)
  if Mode=='Normal': return r
  elif Mode=='Console': return normal2console(r)

# write <style|path> to cell
def cwrite(CellObj, s, Mode='Normal'):
  if Mode=='Normal':  write_style(CellObj, s)
  elif Mode=='Console': write_style(CellObj, console2normal(s))
  CellObj.value=str2data(grab_text(s))


### set or change styles ###
class para_font:
  name='Arial'
  charset=None 
  family=2.0
  b=False
  i=False
  strike=None
  outline=None
  shadow=None
  condense=None
  extend=None
  sz=10.0
  u=None
  vertAlign=None
  scheme=None
  rgb='FF000000' # color
  indexed=0 # color
  auto=True # color
  theme=0 # color
  tint=0 # color
  ftype='rgb' # color

class para_fill:
  patternType=None
  fg_rgb='00000000'
  fg_indexed=0
  fg_auto=True
  fg_theme=0
  fg_tint=0
  fg_type='rgb'
  bg_rgb='00000000'
  bg_indexed=0
  bg_auto=True
  bg_theme=0
  bg_tint=0
  bg_type='rgb'

class para_border:
  outline=True
  diagonalUp=False
  diagonalDown=False
  start=None
  end=None
  ####left####
  lstyle=None

  l_rgb='FF000000'
  l_indexed=0
  l_auto=True
  l_theme=0
  l_tint=0
  l_type='rgb'
  ####right####
  rstyle=None

  r_rgb='FF000000'
  r_indexed=0
  r_auto=True
  r_theme=0
  r_tint=0
  r_type='rgb'
  ####top####
  tstyle=None

  t_rgb='FF000000'
  t_indexed=0
  t_auto=True
  t_theme=0
  t_tint=0
  t_type='rgb'
  ####bottom####
  bstyle=None

  b_rgb='FF000000'
  b_indexed=0
  b_auto=True
  b_theme=0
  b_tint=0
  b_type='rgb'
  ####diagonal####
  dstyle=None

  d_rgb='FF000000'
  d_indexed=0
  d_auto=True
  d_theme=0
  d_tint=0
  d_type='rgb'

  dvertical=None
  dhorizontal=None

  ####veritcal####
  vstyle=None

  v_rgb='FF000000'
  v_indexed=0
  v_auto=True
  v_theme=0
  v_tint=0.0
  v_type='rgb'

  ####horizontal####
  hstyle=None

  h_rgb='FF000000'
  h_indexed=0
  h_auto=True
  h_theme=0
  h_tint=0.0
  h_type='rgb'

class para_alignment:
  horizontal=None
  vertical=None
  textRotation=0
  wrapText=None
  shrinkToFit=None
  indent=0.0
  relativeIndent=0.0
  justifyLastLine=None
  readingOrder=0.0

class para_protection:
  locked=True
  hidden=False

Format='General'

def get_style(Cell, Style):
  '''
  return style of a cell into para_font, para_fill, para_border etc. classes
  Cell: a/Sheet1/A3
  Style: 'font', 'fill', 'border', 'alignment', 'format', 'protection'
  '''
  w, s, c=wb.split_wbname(Cell)
  sh=wb.shto(wb.compose(w, s))
  if Style=='font': 
    f=sh[c].font
    #if type(f.name) is not str: f.name='Calibri'
    #if type(f.charset) is not int: f.charset=0
    #if type(f.family) is not float: f.family=0
    #if type(f.b) is not bool: f.b=False
    #if type(f.i) is not bool: f.i=False
    #if type(f.strike) is not bool: f.strike=False
    #if type(f.outline) is not bool: f.outline=False
    #if type(f.shadow) is not bool: f.shadow=False
    #if type(f.condense) is not bool: f.condense=False
    #if type(f.color) is not xl.styles.colors.Color: f.color=None
    #if type(f.extend) is not bool: f.extend=False
    #if type(f.sz) is not float: f.sz=0
    #if type(f.u) is not bool: f.u=False
    r=para_font()
    r.name=f.name
    r.charset=f.charset
    r.family=f.family
    r.b=f.b
    r.i=f.i
    r.strike=f.strike
    r.outline=f.outline
    r.shadow=f.shadow
    r.condense=f.condense
    
    if f.color:
      r.rgb=f.color.rgb
      r.indexed=f.color.indexed
      r.auto=f.color.auto
      r.theme=f.color.theme
      r.tint=f.color.tint
      r.ftype=f.color.type

    else:
      r.rgb='FF000000'
      r.indexed=0
      r.auto=True
      r.theme=0
      r.tint=0
      r.ftype='rgb'

    r.extend=f.extend
    r.sz=f.sz
    r.u=f.u
    r.vertAlign=f.vertAlign
    r.scheme=f.scheme
    return r
  elif Style=='fill': 
    f=sh[c].fill
    fgC=f.fgColor
    bgC=f.bgColor
    if type(fgC.indexed) is not int: fgC.indexed=0
    if type(fgC.auto) is not bool: fgC.auto=True
    if type(fgC.theme) is not int: fgC.theme=0
    if type(bgC.indexed) is not int: bgC.indexed=0
    if type(bgC.auto) is not bool: bgC.auto=True
    if type(bgC.theme) is not int: bgC.theme=0

    r=para_fill()
    r.patternType=f.patternType
    r.fg_rgb=fgC.rgb
    r.fg_indexed=fgC.indexed
    r.fg_auto=fgC.auto
    r.fg_theme= fgC.theme
    r.fg_tint= fgC.tint
    r.fg_type= fgC.type
    r.bg_rgb= bgC.rgb
    r.bg_indexed= bgC.indexed
    r.bg_auto= bgC.auto
    r.bg_theme= bgC.theme
    r.bg_tint= bgC.tint
    r.bg_type= bgC.type

    return r
 
  elif Style=='border': 
    b=sh[c].border
    r=para_border()

    r.outline=b.outline
    r.diagonalUp=b.diagonalUp
    r.diagonalDown=b.diagonalDown
    r.start=b.start
    r.end=b.end
    
    if b.left:
      r.lstyle=b.left.style
      if b.left.color:
        r.l_rgb=b.left.color.rgb
        r.l_indexed=b.left.color.indexed
        r.l_auto=b.left.color.auto
        r.l_theme=b.left.color.theme
        r.l_tint=b.left.color.tint
        r.l_type=b.left.color.type
      else:
        r.l_rgb='FF000000'
        r.l_indexed=0
        r.l_auto=True
        r.l_theme=0
        r.l_tint=0
        r.l_type='rgb'
    else:
      r.lstyle=None
      r.l_rgb='FF000000'
      r.l_indexed=0
      r.l_auto=True
      r.l_theme=0
      r.l_tint=0
      r.l_type='rgb'

    if b.right:
      r.rstyle=b.right.style
      if b.right.color:
        r.r_rgb=b.right.color.rgb
        r.r_indexed=b.right.color.indexed
        r.r_auto=b.right.color.auto
        r.r_theme=b.right.color.theme
        r.r_tint=b.right.color.tint
        r.r_type=b.right.color.type
      else:
        r.r_rgb='FF000000'
        r.r_indexed=0
        r.r_auto=True
        r.r_theme=0
        r.r_tint=0
        r.r_type='rgb'
    else:
      r.rstyle=None
      r.r_rgb='FF000000'
      r.r_indexed=0
      r.r_auto=True
      r.r_theme=0
      r.r_tint=0
      r.r_type='rgb'

    if b.top:
      r.tstyle=b.top.style
      if b.top.color:
        r.t_rgb=b.top.color.rgb
        r.t_indexed=b.top.color.indexed
        r.t_auto=b.top.color.auto
        r.t_theme=b.top.color.theme
        r.t_tint=b.top.color.tint
        r.t_type=b.top.color.type
      else:
        r.t_rgb='FF000000'
        r.t_indexed=0
        r.t_auto=True
        r.t_theme=0
        r.t_tint=0
        r.t_type='rgb'
    else:
      r.tstyle=None
      r.t_rgb='FF000000'
      r.t_indexed=0
      r.t_auto=True
      r.t_theme=0
      r.t_tint=0
      r.t_type='rgb'

    if b.bottom:
      r.bstyle=b.bottom.style
      if b.bottom.color:
        r.b_rgb=b.bottom.color.rgb
        r.b_indexed=b.bottom.color.indexed
        r.b_auto=b.bottom.color.auto
        r.b_theme=b.bottom.color.theme
        r.b_tint=b.bottom.color.tint
        r.b_type=b.bottom.color.type
      else:
        r.b_rgb='FF000000'
        r.b_indexed=0
        r.b_auto=True
        r.b_theme=0
        r.b_tint=0
        r.b_type='rgb'
    else:
      r.bstyle=None
      r.b_rgb='FF000000'
      r.b_indexed=0
      r.b_auto=True
      r.b_theme=0
      r.b_tint=0
      r.b_type='rgb'
    if b.vertical:
      r.vstyle=b.vertical.style
      if b.vertical.color:
        r.v_rgb=b.vertical.color.rgb
        r.v_indexed=b.vertical.color.indexed
        r.v_auto=b.vertical.color.auto
        r.v_theme=b.vertical.color.theme
        r.v_tint=b.vertical.color.tint
        r.v_type=b.vertical.color.type
      else:
        r.v_rgb='FF000000'
        r.v_indexed=0
        r.v_auto=True
        r.v_theme=0
        r.v_tint=0
        r.v_type='rgb'
    else:
      r.vstyle=None
      r.v_rgb='FF000000'
      r.v_indexed=0
      r.v_auto=True
      r.v_theme=0
      r.v_tint=0
      r.v_type='rgb'
    
    if b.horizontal:
      r.hstyle=b.horizontal.style
      if b.horizontal.color:
        r.h_rgb=b.horizontal.color.rgb
        r.h_indexed=b.horizontal.color.indexed
        r.h_auto=b.horizontal.color.auto
        r.h_theme=b.horizontal.color.theme
        r.h_tint=b.horizontal.color.tint
        r.h_type=b.horizontal.color.type
      else:
        r.h_rgb='FF000000'
        r.h_indexed=0
        r.h_auto=True
        r.h_theme=0
        r.h_tint=0
        r.h_type='rgb'
    else:
      r.hstyle=None
      r.h_rgb='FF000000'
      r.h_indexed=0
      r.h_auto=True
      r.h_theme=0
      r.h_tint=0
      r.h_type='rgb'

    if b.diagonal:
      r.dstyle=b.diagonal.style
      if b.diagonal.color:
        r.d_rgb=b.diagonal.color.rgb
        r.d_indexed=b.diagonal.color.indexed
        r.d_auto=b.diagonal.color.auto
        r.d_theme=b.diagonal.color.theme
        r.d_tint=b.diagonal.color.tint
        r.d_type=b.diagonal.color.type
      else:
        r.d_rgb='FF000000'
        r.d_indexed=0
        r.d_auto=True
        r.d_theme=0
        r.d_tint=0
        r.d_type='rgb'
    else:
      r.dstyle=None
      r.d_rgb='FF000000'
      r.d_indexed=0
      r.d_auto=True
      r.d_theme=0
      r.d_tint=0
      r.d_type='rgb'

    r.dvertical=None
    r.dhorizontal=None

    return r


  elif Style=='alignment': 
    a=sh[c].alignment
    r=para_alignment()

    r.horizontal=a.horizontal
    r.vertical=a.vertical
    r.textRotation=a.textRotation
    r.wrapText=a.wrapText
    r.shrinkToFit=a.shrinkToFit
    r.indent=a.indent
    r.relativeIndent=a.relativeIndent
    r.justifyLastLine=a.justifyLastLine
    r.readingOrder=a.readingOrder

    return r

  elif Style=='format': return sh[c].number_format
  elif Style=='protection': 
    p=sh[c].protection
    r=para_protection()
    r.locked=p.locked
    r.hidden=p.hidden
    return r

def set_style(Cell, Style):
  '''
  set style of a cell, Cell=a/Sheet1/A2
  Style is one of para_font, para_fill, para_border etc.
  For classes: para_font, para_fill etc.: para_font() gets the default values of the font style etc.
  '''
  w, s, c=wb.split_wbname(Cell)
  cell=wb.cello(Cell)
  
  if type(Style) is para_font: cell.font=Font(name=Style.name, sz=Style.sz, b=Style.b, i=Style.i, charset=Style.charset, u=Style.u, strike=Style.strike, color=Style.rgb, scheme=Style.scheme, family=Style.family, vertAlign=Style.vertAlign, outline=Style.outline, shadow=Style.shadow, condense=Style.condense, extend=Style.extend)
  
  elif type(Style) is para_fill:  cell.fill=PatternFill(patternType=Style.patternType, fgColor=Style.fg_rgb, bgColor=Style.bg_rgb)

  elif type(Style) is para_border: cell.border=Border(left=Side(style=Style.lstyle, color=Style.l_rgb), top=Side(style=Style.tstyle, color=Style.t_rgb), bottom=Side(style=Style.bstyle, color=Style.b_rgb), diagonal=Side(style=Style.dstyle, color=Style.d_rgb), vertical=Side(style=Style.vstyle, color=Style.v_rgb), horizontal=Side(style=Style.hstyle, color=Style.h_rgb), diagonalUp=Style.diagonalUp, diagonalDown=Style.diagonalDown, outline=Style.outline, start=Style.start, end=Style.end)
  
  elif type(Style) is para_alignment: cell.alignment=Alignment(horizontal=Style.horizontal, vertical=Style.vertical, textRotation=Style.textRotation, wrapText=Style.wrapText, shrinkToFit=Style.shrinkToFit, indent=Style.indent, relativeIndent=Style.relativeIndent, justifyLastLine=Style.justifyLastLine, readingOrder=Style.readingOrder)
  
  elif type(Style) is para_protection: cell.protection=Protection(locked=Style.locked, hidden=Style.hidden)
  
  elif type(Style) is str: cell.number_format=Style

def fname(Cell, Name=None):
  '''
  set up font name: str
  if no Name, get font name of a cell
  '''
  r=get_style(Cell, 'font')
  if Name:
    r.name=str(Name)
    set_style(Cell, r)
    return r.name
  else: return r.name

def fbold(Cell, Bold=None):
  '''
  set up font bold: bool
  if no bold, get font bold of a cell
  '''
  r=get_style(Cell, 'font')
  if Bold is None: return r.b
  else:
    r.b=Bold
    set_style(Cell, r)
    return r.b
  
def fitalic(Cell, Italic=None):
  '''
  set up font italic: bool
  if no italic, get font italic of a cell
  '''
  r=get_style(Cell, 'font')
  if Italic is None: return r.i
  else:
    r.i=Italic
    set_style(Cell, r)
    return r.i

def fstrike(Cell, Strike=None):
  '''
  set up font strike-through: bool
  if no strike, get font strike-through of a cell
  '''
  r=get_style(Cell, 'font')
  if Strike is None: return r.strike
  else:
    r.strike=Strike
    set_style(Cell, r)
    return r.strike

def foutline(Cell, Outline=None):
  '''
  set up font outline: bool
  if no outline, get font outline of a cell
  '''
  r=get_style(Cell, 'font')
  if Outline is None: return r.outline
  else:
    r.outline=Outline
    set_style(Cell, r)
    return r.outline

def fshadow(Cell, Shadow=None):
  '''
  set up font shadow: bool
  if no shadow, get font shadow of a cell
  '''
  r=get_style(Cell, 'font')
  if Shadow is None: return r.shadow
  else:
    r.shadow=Shadow
    set_style(Cell, r)
    return r.shadow

def fcondense(Cell, Condense=None):
  '''
  set up font condense: bool
  if no condense, get font condense of a cell
  '''
  r=get_style(Cell, 'font')
  if Condense is None: return r.condense
  else:
    r.condense=Condense
    set_style(Cell, r)
    return r.condense   

def fextend(Cell, Extend=None):
  '''
  set up font extend: bool
  if no extend, get font extend of a cell
  '''
  r=get_style(Cell, 'font')
  if Extend is None: return r.extend
  else:
    r.extend=Extend
    set_style(Cell, r)
    return r.extend
  
def fsize(Cell, Size=None):
  '''
  set up font size: float
  if no size, get font size of a cell
  '''
  r=get_style(Cell, 'font')
  if Size is None: return r.sz
  else:
    r.sz=float(Size)
    set_style(Cell, r)
    return r.sz

def funderline(Cell, Underline=None):
  '''
  set up font underline: {'doubleAccounting', 'double', 'single', 'singleAccounting', ''}
  if no underline, get font underline of a cell
  '''
  r=get_style(Cell, 'font')
  if Underline is None: return r.u
  else:
    if Underline: r.u=Underline
    else: r.u=None
    set_style(Cell, r)
    return r.u
    
  
def fvalign(Cell, VertAlign=None):
  '''
  set up font VertAlign: {‘subscript’, ‘superscript’, ‘baseline’, ''}
  if no VertAlign, get font VertAlign of a cell
  '''
  r=get_style(Cell, 'font')
  if VertAlign is None: return r.vertAlign
  else:
    if VertAlign: r.vertAlign=VertAlign
    else: r.vertAlign=None
    set_style(Cell, r)
    return r.vertAlign

def fscheme(Cell, Scheme=None):
  '''
  set up font scheme: {‘major’, ‘minor’, ''}
  if no scheme, get font scheme of a cell
  '''
  r=get_style(Cell, 'font')
  if Scheme is None: return r.scheme
  else:
    if Scheme: r.scheme=Scheme
    else: r.scheme=None
    set_style(Cell, r)
    return r.scheme
  
def fcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up font color
  if no Color, get font corlor of RGB of a cell
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'font')
  if Color is None: return r.rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.rgb=short2rgb(Color, transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.rgb=Color[2:] # RGB color
      else: r.rgb=short2rgb(str(Name2Color[Color]), transparency=Transpanrency) # color name

    set_style(Cell, r)
    return r.rgb

def bgcolor(Cell, Color=None, Fill='solid', Transpanrency=1):
  '''
  set up cell color
  if no Color, get cell corlor of RGB of a cell
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  Fill is the fill type in one of {‘lightGray’, ‘lightDown’, ‘darkUp’, ‘mediumGray’, ‘darkGray’, ‘solid’, ‘lightGrid’, ‘darkHorizontal’, ‘darkTrellis’, ‘darkVertical’, ‘darkDown’, ‘lightHorizontal’, ‘gray125’, ‘darkGrid’, ‘gray0625’, ‘lightTrellis’, ‘lightVertical’, ‘lightUp’}
  '''
  r=get_style(Cell, 'fill')
  if Color is None: return (r.bg_rgb, r.patternType)
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.bg_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.bg_rgb=Color[2:] # RGB color
      else: r.bg_rgb=short2rgb(str(Name2Color[Color]), transparency=Transpanrency) # color name
    r.fg_rgb=r.bg_rgb=r.bg_rgb.upper()
    r.patternType=Fill
    set_style(Cell, r)
    return (r.bg_rgb, r.patternType)
  
def lbstyle(Cell, Style=None):
  '''
  set up left border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.lstyle
  else:
    if Style: r.lstyle=Style
    else: r.lstyle=None
    set_style(Cell, r)
    return r.lstyle

def lbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up left border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.l_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.l_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.l_rgb=Color[2:].upper() # RGB color
      else: r.l_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.l_rgb

def rbstyle(Cell, Style=None):
  '''
  set up right border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.rstyle
  else:
    if Style: r.rstyle=Style
    else: r.rstyle=None
    set_style(Cell, r)
    return r.rstyle

def rbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up right border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.r_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.r_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.r_rgb=Color[2:].upper() # RGB color
      else: r.r_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.r_rgb

def tbstyle(Cell, Style=None):
  '''
  set up top border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.tstyle
  else:
    if Style: r.tstyle=Style
    else: r.tstyle=None
    set_style(Cell, r)
    return r.tstyle

def tbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up top border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.t_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.t_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.t_rgb=Color[2:].upper() # RGB color
      else: r.t_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.t_rgb

def bbstyle(Cell, Style=None):
  '''
  set up bottom border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.bstyle
  else:
    if Style: r.bstyle=Style
    else: r.bstyle=None
    set_style(Cell, r)
    return r.bstyle

def bbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up bottom border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.b_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.b_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.b_rgb=Color[2:].upper() # RGB color
      else: r.b_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.b_rgb

def dbstyle(Cell, Style=None):
  '''
  set up diagonal border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.dstyle
  else:
    if Style: r.dstyle=Style
    else: r.dstyle=None
    set_style(Cell, r)
    return r.dstyle

def dbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up diagonal border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.d_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.d_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.d_rgb=Color[2:].upper() # RGB color
      else: r.d_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.d_rgb

def vbstyle(Cell, Style=None):
  '''
  set up vertical border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.vstyle
  else:
    if Style: r.vstyle=Style
    else: r.vstyle=None
    set_style(Cell, r)
    return r.vstyle

def vbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up vertical border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.v_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.v_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.v_rgb=Color[2:].upper() # RGB color
      else: r.v_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.v_rgb

def hbstyle(Cell, Style=None):
  '''
  set up horizontal border style of a cell: {‘mediumDashDot’, ‘mediumDashed’, ‘dotted’, ‘medium’, ‘thick’, ‘thin’, ‘double’, ‘dashed’, ‘slantDashDot’, ‘dashDot’, ‘dashDotDot’, ‘hair’, ‘mediumDashDotDot’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.hstyle
  else:
    if Style: r.hstyle=Style
    else: r.hstyle=None
    set_style(Cell, r)
    return r.hstyle

def hbcolor(Cell, Color=None, Transpanrency=1):
  '''
  set up horizontal border color of a cell
  if no Color, get cell corlor of RGB of the border
  Color=int: 256 color, str: '0xFF000000' -> rgb color, str: Green: 256 color by name
  '''
  r=get_style(Cell, 'border')
  if Color is None: return r.h_rgb
  else:
    if type(Color) is int: 
      Color=str(Color)
      r.h_rgb=short2rgb(Color,transparency=Transpanrency)
    else:
      Color=str(Color)
      if Color.startswith('0x') or Color.startswith('0X'): r.h_rgb=Color[2:].upper() # RGB color
      else: r.h_rgb=short2rgb(str(Name2Color[Color]).upper(), transparency=Transpanrency) # color name
    set_style(Cell, r)
    return r.h_rgb

def diagup(Cell, Style=None):
  '''
  set up diagonal up: bool
  if no Style, get cell Style of diagonal
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.diagonalUp
  else:
    r.diagonalUp=Style
    set_style(Cell, r)
    return r.diagonalUp

def diagdown(Cell, Style=None):
  '''
  set up diagonal down: bool
  if no Style, get cell Style of diagonal
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.diagonalDown
  else:
    r.diagonalDown=Style
    set_style(Cell, r)
    return r.diagonalDown

def boutline(Cell, Style=None):
  '''
  set up outline: bool
  if no Style, get cell Style of outline
  '''
  r=get_style(Cell, 'border')
  if Style is None: return r.outline
  else:
    r.outline=Style
    set_style(Cell, r)
    return r.outline

def ahorizontal(Cell, Style=None):
  '''
  set up horizontal alignment of a cell: {‘left’, ‘centerContinuous’, ‘center’, ‘distributed’, ‘fill’, ‘justify’, ‘right’, ‘general’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.horizontal
  else:
    if Style: r.horizontal=Style
    else: r.horizontal=None
    set_style(Cell, r)
    return r.horizontal

def avertical(Cell, Style=None):
  '''
  set up horizontal alignment of a cell: {‘bottom’, ‘center’, ‘distributed’, ‘justify’, ‘top’, ''}
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.vertical
  else:
    if Style: r.vertical=Style
    else: r.vertical=None
    set_style(Cell, r)
    return r.vertical

def arotation(Cell, Style=None):
  '''
  set up horizontal alignment of a cell: int: 0-180
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.vertical
  else:
    Style=int(Style)
    if 0<=Style and Style<=180: r.textRotation=Style
    set_style(Cell, r)
    return r.textRotation

def awrap(Cell, Style=None):
  '''
  set up text wrap: bool
  if no Style, get cell Style of text wrap
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.wrapText
  else:
    r.wrapText=Style
    set_style(Cell, r)
    return r.wrapText

def ashrink(Cell, Style=None):
  '''
  set up text shrink to fit the cell: bool
  if no Style, get cell Style of text shrink to fit the cell
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.shrinkToFit
  else:
    r.shrinkToFit=Style
    set_style(Cell, r)
    return r.shrinkToFit

def aindent(Cell, Style=None):
  '''
  set up text indent of a cell: float
  if no Style, get indent of the cell
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.indent
  else:
    r.indent=float(Style)
    set_style(Cell, r)
    return r.indent

def arindent(Cell, Style=None):
  '''
  set up text relative indent of a cell: float
  if no Style, get relative indent of the cell
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.relativeIndent
  else:
    r.relativeIndent=float(Style)
    set_style(Cell, r)
    return r.relativeIndent

def ajlast(Cell, Style=None):
  '''
  set up text line justification: bool
  if no Style, get cell Style of text line justification
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.justifyLastLine
  else:
    r.justifyLastLine=Style
    set_style(Cell, r)
    return r.justifyLastLine

def arorder(Cell, Style=None):
  '''
  set up text reading order of a cell: float
  if no Style, get reading order of the cell
  '''
  r=get_style(Cell, 'alignment')
  if Style is None: return r.readingOrder
  else:
    r.readingOrder=float(Style)
    set_style(Cell, r)
    return r.readingOrder

def plock(Cell, Style=None):
  '''
  set up protection locked: bool
  if no Style, get locked status of protection
  '''
  r=get_style(Cell, 'protection')
  if Style is None: return r.locked
  else:
    r.locked=Style
    set_style(Cell, r)
    return r.locked

def phide(Cell, Style=None):
  '''
  set up protection hidden: bool
  if no Style, get hidden status of protection
  '''
  r=get_style(Cell, 'protection')
  if Style is None: return r.hidden
  else:
    r.hidden=Style
    set_style(Cell, r)
    return r.hidden
  
def format(Cell, Style=None):
  '''
  set up data format (number_format) of a cell
  if no Style, get cell Style of the cell
  if Style=='': set up Style as None
  '''
  r=get_style(Cell, 'format')
  if Style is None: return r
  else:
    set_style(Cell, Style)
    return Style
######## resume here
def clear(Cell, Style=None):
  '''
  clear style of a cell
  if no Style, do nothing
  Style: 'font', 'fill', 'border', 'alignment', 'format', 'protection'
  if Style=='': clear all styles
  '''
  
  if Style is not None:
    if Style=='font': set_style(Cell, para_font())
    elif Style=='fill': set_style(Cell, para_fill())
    elif Style=='border': set_style(Cell, para_border())
    elif Style=='alignment': set_style(Cell, para_alignment())
    elif Style=='format': set_style(Cell, Format)
    elif Style=='protection': set_style(Cell, para_protection())
    elif not Style: 
      for i in (para_font(), para_fill(), para_border(), para_alignment(), Format, para_protection()): set_style(Cell, i)
      
    return Style